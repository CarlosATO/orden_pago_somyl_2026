"""
Módulo independiente para exportación a Excel del Informe de Órdenes de Pago
- Respeta todos los filtros aplicados en el frontend
- Columnas idénticas a las mostradas en la interfaz web
- Datos consistentes con lo que ve el usuario
- Totalmente aislado del módulo principal
"""

from flask import Blueprint, request, current_app, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from app.modules.usuarios import require_modulo
from app.utils.static_data import get_cached_proyectos_with_id
from utils.logger import registrar_log_actividad

bp_excel = Blueprint(
    "excel_informe_pagos", __name__,
    url_prefix="/excel"
)

def get_all_filtered_records(filtros):
    """
    Obtener TODOS los registros que coincidan con los filtros aplicados
    Usa la misma lógica que el módulo principal para garantizar consistencia
    """
    supabase = current_app.config["SUPABASE"]
    
    print(f"[EXCEL] Iniciando exportación con filtros: {filtros}")
    
    try:
        # PASO 1: Obtener todas las órdenes con filtros aplicados (SIN LÍMITES)
        ordenes_query = supabase.table("orden_de_pago").select("orden_numero", count="exact")
        
        if filtros:
            # Filtro por proveedor (usar proveedor_nombre que existe en BD)
            search = filtros.get("proveedor", "").strip()
            if search:
                ordenes_query = ordenes_query.ilike("proveedor_nombre", f"%{search}%")
                print(f"[EXCEL] Filtro proveedor aplicado: {search}")
            
            # Filtro por proyecto
            proyecto_nombre = filtros.get("proyecto", "").strip()
            if proyecto_nombre:
                try:
                    projs = get_cached_proyectos_with_id()
                    for pr in projs:
                        if pr["proyecto"] == proyecto_nombre:
                            ordenes_query = ordenes_query.eq("proyecto", pr["id"])
                            print(f"[EXCEL] Filtro proyecto aplicado: {proyecto_nombre} (ID: {pr['id']})")
                            break
                except Exception as e:
                    print(f"[EXCEL] Error al filtrar por proyecto: {e}")
            
            # Filtros por fecha
            fecha_desde = filtros.get("fecha_desde", "").strip()
            if fecha_desde:
                ordenes_query = ordenes_query.gte("fecha", fecha_desde)
                print(f"[EXCEL] Filtro fecha desde: {fecha_desde}")
                
            fecha_hasta = filtros.get("fecha_hasta", "").strip()
            if fecha_hasta:
                ordenes_query = ordenes_query.lte("fecha", fecha_hasta)
                print(f"[EXCEL] Filtro fecha hasta: {fecha_hasta}")
        
        # OBTENER TODAS LAS ÓRDENES CON PAGINACIÓN FORZADA (SUPABASE LIMIT ~1000)
        print(f"[EXCEL] 🚀 Iniciando obtención completa de órdenes con paginación forzada")
        
        # Primero obtener el conteo total - CORREGIDO: count va en select()
        count_query = supabase.table("orden_de_pago").select("orden_numero", count="exact")
        
        # Aplicar los mismos filtros para el conteo
        if filtros:
            search = filtros.get("proveedor", "").strip()
            if search:
                count_query = count_query.ilike("proveedor_nombre", f"%{search}%")
            
            proyecto_nombre = filtros.get("proyecto", "").strip()
            if proyecto_nombre:
                try:
                    projs = get_cached_proyectos_with_id()
                    for pr in projs:
                        if pr["proyecto"] == proyecto_nombre:
                            count_query = count_query.eq("proyecto", pr["id"])
                            break
                except Exception as e:
                    print(f"[EXCEL] ⚠️ Error aplicando filtro proyecto para conteo: {e}")
            
            fecha_desde = filtros.get("fecha_desde", "").strip()
            if fecha_desde:
                count_query = count_query.gte("fecha", fecha_desde)
                
            fecha_hasta = filtros.get("fecha_hasta", "").strip()
            if fecha_hasta:
                count_query = count_query.lte("fecha", fecha_hasta)
        
        count_result = count_query.execute()
        total_real = count_result.count or 0
        print(f"[EXCEL] 📊 Total de registros que coinciden con filtros: {total_real}")
        
        # SIEMPRE usar paginación por lotes (Supabase limita a ~1000 por consulta)
        all_ordenes = []
        batch_size = 1000
        num_lotes = (total_real + batch_size - 1) // batch_size
        print(f"[EXCEL] 📦 Dividiendo en {num_lotes} lotes de {batch_size} registros")
        
        for lote in range(num_lotes):
            offset = lote * batch_size
            end_range = min(offset + batch_size - 1, total_real - 1)
            
            print(f"[EXCEL] 🔄 Procesando lote {lote + 1}/{num_lotes} (registros {offset}-{end_range})")
            
            # Crear query base para este lote
            batch_query = supabase.table("orden_de_pago").select("orden_numero")
            
            # Reaplicar TODOS los filtros (esto es crucial)
            if filtros:
                search = filtros.get("proveedor", "").strip()
                if search:
                    batch_query = batch_query.ilike("proveedor_nombre", f"%{search}%")
                
                proyecto_nombre = filtros.get("proyecto", "").strip()
                if proyecto_nombre:
                    try:
                        projs = get_cached_proyectos_with_id()
                        for pr in projs:
                            if pr["proyecto"] == proyecto_nombre:
                                batch_query = batch_query.eq("proyecto", pr["id"])
                                break
                    except Exception as e:
                        print(f"[EXCEL] ⚠️ Error aplicando filtro proyecto en lote: {e}")
                
                fecha_desde = filtros.get("fecha_desde", "").strip()
                if fecha_desde:
                    batch_query = batch_query.gte("fecha", fecha_desde)
                    
                fecha_hasta = filtros.get("fecha_hasta", "").strip()
                if fecha_hasta:
                    batch_query = batch_query.lte("fecha", fecha_hasta)
            
            # Ejecutar con range para este lote
            try:
                batch_result = batch_query.range(offset, end_range).execute()
                batch_data = batch_result.data or []
                all_ordenes.extend(batch_data)
                print(f"[EXCEL] ✅ Lote {lote + 1}: {len(batch_data)} registros obtenidos")
            except Exception as e:
                print(f"[EXCEL] ❌ Error en lote {lote + 1}: {e}")
        
        print(f"[EXCEL] 🎯 Total registros obtenidos: {len(all_ordenes)}")
        ordenes_numeros = [r["orden_numero"] for r in all_ordenes]
        
        # Verificar órdenes únicas
        ordenes_unicas = list(set(ordenes_numeros))
        duplicados = len(ordenes_numeros) - len(ordenes_unicas)
        print(f"[EXCEL] 📋 Órdenes únicas: {len(ordenes_unicas)} (duplicados: {duplicados})")
        
        ordenes_numeros = ordenes_unicas  # Usar solo las únicas
        
        if not ordenes_numeros:
            print("[EXCEL] No se encontraron órdenes con los filtros aplicados")
            return []
        
        print(f"[EXCEL] Encontradas {len(ordenes_numeros)} órdenes únicas")
        
        # PASO 2: Obtener datos completos por lotes (SIN LÍMITES)
        all_rows = []
        ordenes_encontradas_set = set()
        batch_size = 1000  # Aumentar tamaño del lote
        for i in range(0, len(ordenes_numeros), batch_size):
            batch_numeros = ordenes_numeros[i:i+batch_size]
            print(f"[EXCEL] 🔍 Lote {i//batch_size + 1}: Buscando datos para {len(batch_numeros)} órdenes")
            print(f"[EXCEL] 📋 Órdenes solicitadas (muestra): {batch_numeros[:5]}...")
            
            batch_data = supabase.table("orden_de_pago").select(
                "orden_numero, fecha, proveedor, proveedor_nombre, detalle_compra, "
                "factura, proyecto, orden_compra, condicion_pago, "
                "vencimiento, fecha_factura, costo_final_con_iva, ingreso_id"
            ).in_("orden_numero", batch_numeros).execute().data or []
            
            # Verificar qué órdenes se encontraron
            ordenes_en_batch = set(r["orden_numero"] for r in batch_data)
            ordenes_encontradas_set.update(ordenes_en_batch)
            ordenes_faltantes = set(batch_numeros) - ordenes_en_batch
            
            print(f"[EXCEL] ✅ Lote {i//batch_size + 1}: {len(batch_data)} registros obtenidos para {len(ordenes_en_batch)} órdenes")
            if ordenes_faltantes:
                print(f"[EXCEL] ⚠️ Lote {i//batch_size + 1}: {len(ordenes_faltantes)} órdenes SIN DATOS: {list(ordenes_faltantes)[:10]}...")
            
            all_rows.extend(batch_data)
        
        ordenes_sin_datos = set(ordenes_numeros) - ordenes_encontradas_set
        print(f"[EXCEL] 📊 RESUMEN:")
        print(f"[EXCEL] 📋 Órdenes solicitadas: {len(ordenes_numeros)}")
        print(f"[EXCEL] ✅ Órdenes con datos: {len(ordenes_encontradas_set)}")
        print(f"[EXCEL] ❌ Órdenes sin datos: {len(ordenes_sin_datos)}")
        if ordenes_sin_datos:
            print(f"[EXCEL] 🔍 Ejemplos de órdenes sin datos: {list(ordenes_sin_datos)[:20]}")
        print(f"[EXCEL] 📄 Total registros de datos obtenidos: {len(all_rows)}")
        
        # PASO 3: Enriquecer con nombres de proyectos
        try:
            projs = get_cached_proyectos_with_id()
            proyecto_map = {pr["id"]: pr["proyecto"] for pr in projs}
            for r in all_rows:
                proj_id = r.get("proyecto")
                r["proyecto_nombre"] = proyecto_map.get(proj_id, f"ID:{proj_id}" if proj_id else "Sin Proyecto")
        except Exception as e:
            print(f"[EXCEL] Error al enriquecer proyectos: {e}")
        
        # PASO 4: Agrupar por orden_numero (misma lógica que el frontend)
        pagos_dict = {}
        for r in all_rows:
            num = r["orden_numero"]
            if num not in pagos_dict:
                pagos_dict[num] = {
                    "orden_numero": num,
                    "fecha": r["fecha"],
                    "proveedor": r.get("proveedor"),
                    "proveedor_nombre": r.get("proveedor_nombre", r.get("proveedor", "")),  # CORREGIDO
                    "rut_proveedor": r.get("rut_proveedor", "-"),  # AGREGADO
                    "detalle_compra": r["detalle_compra"],
                    "factura": r["factura"],
                    "total_pago": 0.0,
                    "proyecto": r.get("proyecto_nombre", "Sin Proyecto"),
                    "item": r.get("item", "-"),  # AGREGADO
                    "orden_compra": r["orden_compra"],
                    "condicion_pago": r["condicion_pago"],
                    "vencimiento": r["vencimiento"],
                    "fecha_factura": r["fecha_factura"],
                    "fecha_pago": None,
                    "total_abonado": 0.0,
                    "saldo_pendiente": 0.0,
                    "estado": "Pendiente"  # Se calculará después
                }
            pagos_dict[num]["total_pago"] += float(r.get("costo_final_con_iva") or 0)
        
        print(f"[EXCEL] Órdenes agrupadas: {len(pagos_dict)}")
        
        # PASO 5: Obtener fechas de pago
        try:
            batch_size_fechas = 1000
            fecha_map = {}
            for i in range(0, len(ordenes_numeros), batch_size_fechas):
                batch_nums = ordenes_numeros[i:i+batch_size_fechas]
                fechas_batch = supabase.table("fechas_de_pagos_op").select("orden_numero, fecha_pago").in_("orden_numero", batch_nums).execute().data or []
                for row in fechas_batch:
                    fecha_map[row["orden_numero"]] = row["fecha_pago"]
            
            for num in pagos_dict:
                pagos_dict[num]["fecha_pago"] = fecha_map.get(num)
            
            print(f"[EXCEL] Fechas de pago cargadas: {len(fecha_map)}")
        except Exception as e:
            print(f"[EXCEL] Error al obtener fechas de pago: {e}")
        
        # PASO 6: Calcular abonos (CORREGIDO)
        try:
            batch_size_abonos = 1000
            abonos_map = {}
            total_abonos_procesados = 0
            
            for i in range(0, len(ordenes_numeros), batch_size_abonos):
                batch_nums = ordenes_numeros[i:i+batch_size_abonos]
                abonos_batch = supabase.table("abonos_op").select("orden_numero, monto_abono").in_("orden_numero", batch_nums).execute().data or []
                
                for row in abonos_batch:
                    num = row["orden_numero"]
                    monto = float(row.get("monto_abono", 0))  # CORREGIDO: usar monto_abono
                    if num not in abonos_map:
                        abonos_map[num] = 0.0
                    abonos_map[num] += monto
                    total_abonos_procesados += 1
            
            # Aplicar abonos y calcular saldos
            for num in pagos_dict:
                total_abonado = abonos_map.get(num, 0.0)
                pagos_dict[num]["total_abonado"] = total_abonado
                pagos_dict[num]["saldo_pendiente"] = max(0, pagos_dict[num]["total_pago"] - total_abonado)
            
            print(f"[EXCEL] Abonos procesados: {total_abonos_procesados} registros para {len(abonos_map)} órdenes con abonos")
        except Exception as e:
            print(f"[EXCEL] Error al calcular abonos: {e}")
        
        # PASO 7: Calcular estado y aplicar filtro de estado (CORREGIDO)
        def calcular_estado(p):
            if p.get("fecha_pago"):
                return "pagado"  # CORREGIDO: usar mismo formato que frontend
            elif p.get("total_abonado", 0) > 0:
                return "abono"   # CORREGIDO: usar mismo formato que frontend
            else:
                return "pendiente"  # CORREGIDO: usar mismo formato que frontend
        
        # Calcular estado para todos los registros
        for num in pagos_dict:
            pagos_dict[num]["estado"] = calcular_estado(pagos_dict[num])
        
        # Convertir a lista ordenada
        pagos_procesados = [pagos_dict[k] for k in sorted(pagos_dict.keys(), reverse=True)]
        
        # Aplicar filtro de estado si existe (CORREGIDO - MAPEO EXACTO CON FRONTEND)
        if filtros and filtros.get("estado"):
            estado_filtro = filtros["estado"].lower().strip()
            
            # Filtrar directamente con estados en minúsculas (mismo formato que frontend)
            pagos_procesados = [p for p in pagos_procesados if p["estado"] == estado_filtro]
            
            print(f"[EXCEL] Filtro de estado '{estado_filtro}' aplicado: {len(pagos_procesados)} registros")
        
        print(f"[EXCEL] ✅ TOTAL REGISTROS FINALES PARA EXPORTAR: {len(pagos_procesados)}")
        return pagos_procesados
        
    except Exception as e:
        print(f"[EXCEL ERROR] Error general en obtención de datos: {e}")
        import traceback
        traceback.print_exc()
        return []

def format_estado_excel(estado):
    """Formatear estado para mostrar en Excel de manera legible"""
    if estado == "pagado":
        return "Pagado"
    elif estado == "pendiente":
        return "Pendiente"
    elif estado == "abono":
        return "Con Abonos"
    else:
        return estado

def format_currency_number(value):
    """Formatear como número entero para Excel (sin $ para que Excel pueda sumar)"""
    try:
        if isinstance(value, (int, float)):
            return int(value)
        return 0
    except:
        return 0

def format_date(date_str):
    """Formatear fecha para Excel"""
    if not date_str:
        return ""
    try:
        if isinstance(date_str, str):
            return date_str
        return str(date_str)
    except:
        return ""

@bp_excel.route("/informe_op", methods=["GET"])
@login_required
@require_modulo('pagos')
def export_informe_op():
    """
    Exportar informe de órdenes de pago a Excel
    Respeta exactamente los mismos filtros que el frontend
    """
    try:
        # Obtener filtros exactos del frontend
        filtros = {}
        proveedor = request.args.get("proveedor", "").strip()
        proyecto = request.args.get("proyecto", "").strip()
        estado = request.args.get("estado", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()

        print(f"[EXCEL EXPORT] Parámetros recibidos:")
        print(f"  - proveedor: '{proveedor}'")
        print(f"  - proyecto: '{proyecto}'")
        print(f"  - estado: '{estado}'")
        print(f"  - fecha_desde: '{fecha_desde}'")
        print(f"  - fecha_hasta: '{fecha_hasta}'")

        if proveedor:
            filtros["proveedor"] = proveedor
        if proyecto:
            filtros["proyecto"] = proyecto
        if estado:
            filtros["estado"] = estado
        if fecha_desde:
            filtros["fecha_desde"] = fecha_desde
        if fecha_hasta:
            filtros["fecha_hasta"] = fecha_hasta

        print(f"[EXCEL EXPORT] Filtros construidos: {filtros}")

        # Obtener datos con filtros aplicados
        pagos = get_all_filtered_records(filtros)
        
        if not pagos:
            # Si no hay datos, crear Excel vacío con mensaje
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe Órdenes de Pago"
            ws.append(["No se encontraron registros con los filtros aplicados"])
        else:
            # Crear workbook con datos
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe Órdenes de Pago"

            # Headers - exactamente las mismas columnas que ve el usuario
            headers = [
                "Orden N°", 
                "Fecha", 
                "Proveedor", 
                "RUT Proveedor",  # AGREGADO
                "Detalle", 
                "Factura", 
                "Total Pago",
                "Total Abonado", 
                "Saldo Pendiente", 
                "Estado",
                "Proyecto", 
                "Items",  # AGREGADO  
                "OC", 
                "Condición Pago",
                "Vencimiento", 
                "Fecha Factura", 
                "Fecha de Pago"
            ]
            
            # Agregar headers con formato
            ws.append(headers)
            
            # Formatear headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Agregar datos - exactamente como se muestran en el frontend
            for p in pagos:
                row = [
                    p.get("orden_numero", ""),
                    format_date(p.get("fecha", "")),
                    p.get("proveedor_nombre", ""),  # CORREGIDO: usar proveedor_nombre en lugar de proveedor
                    p.get("rut_proveedor", "-"),  # AGREGADO
                    p.get("detalle_compra", ""),
                    p.get("factura", ""),
                    format_currency_number(p.get("total_pago", 0)),      # NÚMEROS PLANOS
                    format_currency_number(p.get("total_abonado", 0)),   # NÚMEROS PLANOS
                    format_currency_number(p.get("saldo_pendiente", 0)), # NÚMEROS PLANOS
                    format_estado_excel(p.get("estado", "")),  # ESTADO FORMATEADO
                    p.get("proyecto", ""),
                    p.get("item", "-"),  # AGREGADO
                    p.get("orden_compra", ""),
                    p.get("condicion_pago", ""),
                    format_date(p.get("vencimiento", "")),
                    format_date(p.get("fecha_factura", "")),
                    format_date(p.get("fecha_pago", ""))
                ]
                ws.append(row)

            # Ajustar ancho de columnas automáticamente
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nombre del archivo con timestamp y filtros
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filtros_str = ""
        if filtros:
            filtros_aplicados = []
            if filtros.get("estado"):
                filtros_aplicados.append(f"estado_{filtros['estado']}")
            if filtros.get("proyecto"):
                filtros_aplicados.append(f"proyecto")
            if filtros.get("proveedor"):
                filtros_aplicados.append(f"proveedor")
            if filtros_aplicados:
                filtros_str = f"_{'_'.join(filtros_aplicados)}"
        
        filename = f"informe_ordenes_pago_{timestamp}{filtros_str}.xlsx"

        # Registrar actividad
        registrar_log_actividad("Exportar Excel", f"Informe OP - Filtros: {filtros} - {len(pagos)} registros")

        return send_file(
            output, 
            download_name=filename, 
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"[EXCEL ERROR] Error en exportación: {e}")
        import traceback
        traceback.print_exc()
        
        # En caso de error, devolver Excel con mensaje de error
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws.append([f"Error al generar reporte: {str(e)}"])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output, 
            download_name=f"error_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", 
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
