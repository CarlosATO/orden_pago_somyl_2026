"""
Módulo de gastos directos adaptado para trabajar con frontend React
Incluye funcionalidad de importación/exportación Excel
"""

from flask import Blueprint, request, jsonify, current_app, send_file
from backend.utils.decorators import token_required
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

bp = Blueprint("gastos_directos", __name__)


def validar_mes(mes_valor):
    """
    Valida y convierte el mes a número (1-12)
    Acepta: número (1-12) o nombre del mes en español
    """
    meses_map = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    
    # Si es número
    if isinstance(mes_valor, (int, float)):
        mes = int(mes_valor)
        if 1 <= mes <= 12:
            return mes, None
        return None, f"Mes numérico inválido: {mes}"
    
    # Si es texto
    mes_str = str(mes_valor).lower().strip()
    
    # Intentar convertir a número
    try:
        mes = int(mes_str)
        if 1 <= mes <= 12:
            return mes, None
        return None, f"Mes numérico inválido: {mes}"
    except ValueError:
        pass
    
    # Buscar en el mapa de nombres
    mes = meses_map.get(mes_str)
    if mes:
        return mes, None
    
    return None, f"Mes no reconocido: '{mes_valor}'"


def validar_fecha(fecha_valor):
    """
    Valida y convierte la fecha al formato ISO
    Acepta: datetime, string YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY
    """
    if not fecha_valor:
        return datetime.now().date().isoformat(), None
    
    # Si ya es datetime
    if isinstance(fecha_valor, datetime):
        return fecha_valor.date().isoformat(), None
    
    # Si es string
    fecha_str = str(fecha_valor).strip()
    
    # Intentar varios formatos
    formatos = [
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%d/%m/%Y',
        '%Y/%m/%d'
    ]
    
    for formato in formatos:
        try:
            fecha_obj = datetime.strptime(fecha_str, formato)
            return fecha_obj.date().isoformat(), None
        except ValueError:
            continue
    
    return None, f"Formato de fecha inválido: '{fecha_valor}'"


@bp.route("/todos", methods=["GET"])
@token_required
def api_get_gastos_directos(current_user):
    """
    Devuelve lista de todos los gastos directos en formato JSON
    Opcionalmente filtrados por proyecto
    """
    try:
        supabase = current_app.config['SUPABASE']
        proyecto_id = request.args.get('proyecto_id', type=int)
        
        query = supabase.table("gastos_directos").select("*")
        
        if proyecto_id:
            query = query.eq("proyecto_id", proyecto_id)
        
        query = query.order("fecha", desc=True).order("id", desc=True)
        
        response = query.execute()
        gastos = response.data or []
        
        # Obtener información adicional de items
        if gastos:
            items_ids = list(set([g.get('item_id') for g in gastos if g.get('item_id')]))
            if items_ids:
                items_result = supabase.table("item").select("id, tipo").in_("id", items_ids).execute()
                items_dict = {item['id']: item for item in (items_result.data or [])}
                
                # Agregar info del item a cada gasto
                for gasto in gastos:
                    item_id = gasto.get('item_id')
                    if item_id and item_id in items_dict:
                        gasto['item_nombre'] = items_dict[item_id]['tipo']
        
        return jsonify({
            "success": True,
            "count": len(gastos),
            "data": gastos
        })
        
    except Exception as e:
        current_app.logger.error(f"Error al obtener gastos directos: {str(e)}")
        return jsonify({"success": False, "message": "Error al obtener los gastos"}), 500


@bp.route("/new", methods=["POST"])
@token_required
def new_gasto_directo(current_user):
    """Crea un nuevo gasto directo"""
    try:
        data = request.get_json() or {}
        
        # Validar campos requeridos
        proyecto_id = data.get("proyecto_id")
        item_id = data.get("item_id")
        monto = data.get("monto")
        mes_valor = data.get("mes")
        
        if not all([proyecto_id, item_id, monto, mes_valor]):
            return jsonify({"success": False, "message": "Faltan campos obligatorios"}), 400
        
        # Validar monto
        try:
            monto = float(monto)
            if monto <= 0:
                return jsonify({"success": False, "message": "El monto debe ser mayor a 0"}), 400
        except (ValueError, TypeError):
            return jsonify({"success": False, "message": "Monto inválido"}), 400
        
        # Validar mes
        mes, error_mes = validar_mes(mes_valor)
        if error_mes:
            return jsonify({"success": False, "message": error_mes}), 400
        
        # Validar fecha
        fecha_valor = data.get("fecha")
        fecha, error_fecha = validar_fecha(fecha_valor)
        if error_fecha:
            return jsonify({"success": False, "message": error_fecha}), 400
        
        supabase = current_app.config['SUPABASE']
        
        # Verificar que proyecto existe
        proyecto_check = supabase.table("proyectos").select("id").eq("id", proyecto_id).limit(1).execute()
        if not proyecto_check.data:
            return jsonify({"success": False, "message": "Proyecto no existe"}), 400
        
        # Verificar que item existe
        item_check = supabase.table("item").select("id").eq("id", item_id).limit(1).execute()
        if not item_check.data:
            return jsonify({"success": False, "message": "Item no existe"}), 400
        
        # Preparar datos para insertar
        nuevo_gasto = {
            "proyecto_id": int(proyecto_id),
            "item_id": int(item_id),
            "descripcion": data.get("descripcion", "").strip().upper() or None,
            "mes": mes,
            "monto": int(monto),
            "fecha": fecha,
            "usuario_id": current_user['id']
        }
        
        # Insertar en BD
        result = supabase.table("gastos_directos").insert(nuevo_gasto).execute()
        
        if result.data:
            return jsonify({
                "success": True,
                "message": "Gasto directo creado exitosamente",
                "data": result.data[0]
            })
        else:
            return jsonify({"success": False, "message": "Error al crear el gasto"}), 500
            
    except Exception as e:
        current_app.logger.error(f"Error al crear gasto directo: {str(e)}")
        return jsonify({"success": False, "message": "Error interno del servidor"}), 500


@bp.route("/<int:gasto_id>", methods=["DELETE"])
@token_required
def delete_gasto_directo(current_user, gasto_id):
    """Elimina un gasto directo"""
    try:
        supabase = current_app.config['SUPABASE']
        
        # Verificar que el gasto existe
        gasto_check = supabase.table("gastos_directos").select("id").eq("id", gasto_id).limit(1).execute()
        
        if not gasto_check.data:
            return jsonify({"success": False, "message": "Gasto no encontrado"}), 404
        
        # Eliminar
        result = supabase.table("gastos_directos").delete().eq("id", gasto_id).execute()
        
        if result.data:
            return jsonify({"success": True, "message": "Gasto eliminado exitosamente"})
        else:
            return jsonify({"success": False, "message": "Error al eliminar el gasto"}), 500
            
    except Exception as e:
        current_app.logger.error(f"Error al eliminar gasto: {str(e)}")
        return jsonify({"success": False, "message": "Error interno del servidor"}), 500


@bp.route("/plantilla-excel", methods=["GET"])
@token_required
def descargar_plantilla_excel(current_user):
    """
    Genera y descarga una plantilla Excel para importar gastos directos
    Incluye hojas de referencia con proyectos e items
    """
    try:
        supabase = current_app.config['SUPABASE']
        
        # Crear workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # === HOJA 1: GASTOS (para completar) ===
        ws_gastos = wb.create_sheet("Gastos", 0)
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["proyecto_id", "item_id", "descripcion", "mes", "monto", "fecha"]
        for col, header in enumerate(headers, 1):
            cell = ws_gastos.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Fila de ejemplo
        ejemplo = [
            "1", 
            "5", 
            "MATERIALES DE CONSTRUCCIÓN", 
            "1", 
            "500000", 
            datetime.now().strftime("%Y-%m-%d")
        ]
        for col, valor in enumerate(ejemplo, 1):
            cell = ws_gastos.cell(row=2, column=col, value=valor)
            cell.border = border
        
        # Ajustar anchos
        ws_gastos.column_dimensions['A'].width = 15
        ws_gastos.column_dimensions['B'].width = 10
        ws_gastos.column_dimensions['C'].width = 40
        ws_gastos.column_dimensions['D'].width = 12
        ws_gastos.column_dimensions['E'].width = 15
        ws_gastos.column_dimensions['F'].width = 15
        
        # Nota informativa
        ws_gastos.cell(row=4, column=1, value="INSTRUCCIONES:")
        ws_gastos.cell(row=5, column=1, value="• Consulte las hojas 'Proyectos' e 'Items' para ver los IDs disponibles")
        ws_gastos.cell(row=6, column=1, value="• El mes puede ser número (1-12) o nombre (enero, febrero, etc.)")
        ws_gastos.cell(row=7, column=1, value="• La fecha debe estar en formato YYYY-MM-DD")
        ws_gastos.cell(row=8, column=1, value="• El monto no debe incluir puntos ni comas, solo números")
        
        # === HOJA 2: PROYECTOS (referencia) ===
        ws_proyectos = wb.create_sheet("Proyectos", 1)
        
        # Headers
        ws_proyectos.cell(row=1, column=1, value="ID")
        ws_proyectos.cell(row=1, column=2, value="NOMBRE PROYECTO")
        
        for col in [1, 2]:
            cell = ws_proyectos.cell(row=1, column=col)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Obtener proyectos
        proyectos = supabase.table("proyectos").select("id, proyecto").order("proyecto").execute().data or []
        
        for idx, proyecto in enumerate(proyectos, 2):
            ws_proyectos.cell(row=idx, column=1, value=proyecto['id']).border = border
            ws_proyectos.cell(row=idx, column=2, value=proyecto['proyecto']).border = border
        
        ws_proyectos.column_dimensions['A'].width = 10
        ws_proyectos.column_dimensions['B'].width = 50
        
        # === HOJA 3: ITEMS (referencia) ===
        ws_items = wb.create_sheet("Items", 2)
        
        # Headers
        ws_items.cell(row=1, column=1, value="ID")
        ws_items.cell(row=1, column=2, value="TIPO ITEM")
        
        for col in [1, 2]:
            cell = ws_items.cell(row=1, column=col)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Obtener items
        items = supabase.table("item").select("id, tipo").order("tipo").execute().data or []
        
        for idx, item in enumerate(items, 2):
            ws_items.cell(row=idx, column=1, value=item['id']).border = border
            ws_items.cell(row=idx, column=2, value=item['tipo']).border = border
        
        ws_items.column_dimensions['A'].width = 10
        ws_items.column_dimensions['B'].width = 40
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'plantilla_gastos_directos_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        current_app.logger.error(f"Error al generar plantilla Excel: {str(e)}")
        return jsonify({"success": False, "message": "Error al generar la plantilla"}), 500


@bp.route("/validar-excel", methods=["POST"])
@token_required
def validar_excel(current_user):
    """
    Valida un archivo Excel antes de importarlo
    Retorna preview de datos y errores encontrados
    """
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "No se recibió ningún archivo"}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({"success": False, "message": "Nombre de archivo vacío"}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"success": False, "message": "El archivo debe ser Excel (.xlsx o .xls)"}), 400
        
        # Leer Excel
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb["Gastos"]
        except KeyError:
            return jsonify({"success": False, "message": "El archivo debe contener una hoja llamada 'Gastos'"}), 400
        except Exception as e:
            return jsonify({"success": False, "message": f"Error al leer el archivo: {str(e)}"}), 400
        
        supabase = current_app.config['SUPABASE']
        
        # Obtener proyectos e items válidos
        proyectos = supabase.table("proyectos").select("id").execute().data or []
        items = supabase.table("item").select("id").execute().data or []
        
        proyectos_ids = set([p['id'] for p in proyectos])
        items_ids = set([i['id'] for i in items])
        
        # Procesar filas
        datos_validos = []
        errores = []
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Saltar filas vacías
            if not any(row):
                continue
            
            proyecto_id, item_id, descripcion, mes_valor, monto_valor, fecha_valor = row[:6]
            
            errores_fila = []
            
            # Validar proyecto_id
            try:
                proyecto_id = int(proyecto_id)
                if proyecto_id not in proyectos_ids:
                    errores_fila.append(f"Proyecto ID {proyecto_id} no existe")
            except (ValueError, TypeError):
                errores_fila.append(f"proyecto_id inválido: '{proyecto_id}'")
                proyecto_id = None
            
            # Validar item_id
            try:
                item_id = int(item_id)
                if item_id not in items_ids:
                    errores_fila.append(f"Item ID {item_id} no existe")
            except (ValueError, TypeError):
                errores_fila.append(f"item_id inválido: '{item_id}'")
                item_id = None
            
            # Validar mes
            mes, error_mes = validar_mes(mes_valor)
            if error_mes:
                errores_fila.append(error_mes)
            
            # Validar monto
            try:
                monto = float(monto_valor)
                if monto <= 0:
                    errores_fila.append(f"Monto debe ser mayor a 0")
            except (ValueError, TypeError):
                errores_fila.append(f"Monto inválido: '{monto_valor}'")
                monto = None
            
            # Validar fecha
            fecha, error_fecha = validar_fecha(fecha_valor)
            if error_fecha:
                errores_fila.append(error_fecha)
            
            # Descripción (opcional)
            descripcion = str(descripcion).strip().upper() if descripcion else None
            
            if errores_fila:
                errores.append({
                    "fila": idx,
                    "errores": errores_fila,
                    "datos": {
                        "proyecto_id": proyecto_id,
                        "item_id": item_id,
                        "descripcion": descripcion,
                        "mes": mes_valor,
                        "monto": monto_valor,
                        "fecha": fecha_valor
                    }
                })
            else:
                datos_validos.append({
                    "proyecto_id": proyecto_id,
                    "item_id": item_id,
                    "descripcion": descripcion,
                    "mes": mes,
                    "monto": int(monto),
                    "fecha": fecha
                })
        
        return jsonify({
            "success": True,
            "datos_validos": datos_validos,
            "total_validos": len(datos_validos),
            "errores": errores,
            "total_errores": len(errores)
        })
        
    except Exception as e:
        current_app.logger.error(f"Error al validar Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Error al validar el archivo: {str(e)}"}), 500


@bp.route("/importar-excel", methods=["POST"])
@token_required
def importar_excel(current_user):
    """
    Importa gastos directos desde datos previamente validados
    """
    try:
        data = request.get_json() or {}
        gastos = data.get("gastos", [])
        
        if not gastos:
            return jsonify({"success": False, "message": "No hay gastos para importar"}), 400
        
        supabase = current_app.config['SUPABASE']
        
        # Agregar usuario_id a cada gasto
        for gasto in gastos:
            gasto["usuario_id"] = current_user['id']
        
        # Insertar en lote
        result = supabase.table("gastos_directos").insert(gastos).execute()
        
        if result.data:
            return jsonify({
                "success": True,
                "message": f"Se importaron {len(result.data)} gastos exitosamente",
                "total": len(result.data)
            })
        else:
            return jsonify({"success": False, "message": "Error al importar los gastos"}), 500
            
    except Exception as e:
        current_app.logger.error(f"Error al importar Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Error al importar: {str(e)}"}), 500


@bp.route("/exportar-excel/<int:proyecto_id>", methods=["GET"])
@token_required
def exportar_gastos_excel(current_user, proyecto_id):
    """
    Exporta los gastos directos de un proyecto a Excel
    """
    try:
        supabase = current_app.config['SUPABASE']
        
        # Obtener información del proyecto
        proyecto = supabase.table("proyectos").select("proyecto").eq("id", proyecto_id).limit(1).execute()
        
        if not proyecto.data:
            return jsonify({"success": False, "message": "Proyecto no encontrado"}), 404
        
        nombre_proyecto = proyecto.data[0]['proyecto']
        
        # Obtener gastos del proyecto
        gastos = supabase.table("gastos_directos") \
            .select("*") \
            .eq("proyecto_id", proyecto_id) \
            .order("fecha", desc=True) \
            .execute().data or []
        
        if not gastos:
            return jsonify({"success": False, "message": "No hay gastos para exportar"}), 404
        
        # Obtener nombres de items
        items_ids = list(set([g.get('item_id') for g in gastos if g.get('item_id')]))
        items_result = supabase.table("item").select("id, tipo").in_("id", items_ids).execute()
        items_dict = {item['id']: item['tipo'] for item in (items_result.data or [])}
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos Directos"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"GASTOS DIRECTOS - {nombre_proyecto}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Headers
        headers = ["ID", "Proyecto ID", "Item ID", "Item", "Descripción", "Mes", "Monto", "Fecha"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        
        for idx, gasto in enumerate(gastos, 4):
            mes_num = gasto.get('mes', 0)
            mes_nombre = meses_nombres[mes_num - 1] if 1 <= mes_num <= 12 else str(mes_num)
            
            ws.cell(row=idx, column=1, value=gasto.get('id')).border = border
            ws.cell(row=idx, column=2, value=gasto.get('proyecto_id')).border = border
            ws.cell(row=idx, column=3, value=gasto.get('item_id')).border = border
            ws.cell(row=idx, column=4, value=items_dict.get(gasto.get('item_id'), '')).border = border
            ws.cell(row=idx, column=5, value=gasto.get('descripcion', '')).border = border
            ws.cell(row=idx, column=6, value=mes_nombre).border = border
            ws.cell(row=idx, column=7, value=gasto.get('monto', 0)).border = border
            ws.cell(row=idx, column=8, value=gasto.get('fecha', '')).border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"gastos_{nombre_proyecto.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Error al exportar Excel: {str(e)}")
        return jsonify({"success": False, "message": "Error al exportar los datos"}), 500