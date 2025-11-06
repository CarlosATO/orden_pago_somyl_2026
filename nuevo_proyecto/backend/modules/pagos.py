# nuevo_proyecto/backend/modules/pagos.py

from flask import Blueprint, request, jsonify, current_app, send_file
from datetime import date, datetime, timedelta
from backend.utils.decorators import token_required
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time

bp = Blueprint("pagos_api", __name__)
logger = logging.getLogger(__name__)

# ================================================================
# CACHÉ SIMPLE EN MEMORIA PARA DATOS PROCESADOS
# ================================================================

_pagos_cache = {
    "data": None,
    "timestamp": 0,
    "ttl": 60  # 60 segundos de vida útil
}

def invalidar_cache_pagos():
    """Invalida el caché de pagos (llamar cuando se modifiquen datos)"""
    _pagos_cache["data"] = None
    _pagos_cache["timestamp"] = 0

# ================================================================
# FUNCIONES AUXILIARES
# ================================================================

def get_cached_proveedores():
    """Obtener proveedores desde Supabase"""
    supabase = current_app.config['SUPABASE']
    try:
        result = supabase.table("proveedores").select("id, nombre, rut, cuenta").execute()
        return result.data or []
    except Exception as e:
        logger.error(f"Error obteniendo proveedores: {e}")
        return []

def get_cached_proyectos():
    """Obtener proyectos desde Supabase"""
    supabase = current_app.config['SUPABASE']
    try:
        result = supabase.table("proyectos").select("id, proyecto").execute()
        return result.data or []
    except Exception as e:
        logger.error(f"Error obteniendo proyectos: {e}")
        return []

def calcular_estado_pago(fecha_pago, total_abonado, total_pago):
    """
    Calcular estado del pago de manera simple y clara:
    1. Si tiene fecha_pago -> PAGADO (sin importar abonos)
    2. Si tiene abonos pero no fecha -> ABONO
    3. Si no tiene fecha ni abonos -> PENDIENTE
    """
    # Caso 1: Si tiene fecha de pago, está PAGADO
    if fecha_pago:
        return "pagado"
    
    # Caso 2: Si no tiene fecha pero tiene abonos, está en ABONO
    if (total_abonado or 0) > 0:
        return "abono"
    
    # Caso 3: Sin fecha y sin abonos -> PENDIENTE
    return "pendiente"

def formatear_fecha_chilena(fecha_valor, default="---"):
    """
    Formatea una fecha al formato chileno DD/MM/YYYY.
    Maneja strings en formato ISO y objetos date/datetime.
    Retorna default si la fecha es None o inválida.
    """
    if not fecha_valor:
        return default
    
    try:
        if isinstance(fecha_valor, str):
            # Intentar parsear desde formato ISO YYYY-MM-DD
            fecha_obj = datetime.strptime(fecha_valor, "%Y-%m-%d")
            return fecha_obj.strftime("%d/%m/%Y")
        elif isinstance(fecha_valor, (date, datetime)):
            return fecha_valor.strftime("%d/%m/%Y")
    except Exception:
        pass
    
    return default

def obtener_todos_pagos_procesados(supabase, filtros_base=None):
    """
    Obtiene TODOS los pagos con estados calculados.
    Esta función es intensiva, por eso usamos caché.
    
    filtros_base: dict con filtros que se pueden aplicar en BD (proveedor, proyecto, fechas, orden_numero)
    """
    # Construir query base
    query = supabase.table("orden_de_pago").select(
        "orden_numero, fecha, proveedor, proveedor_nombre, detalle_compra, "
        "factura, costo_final_con_iva, proyecto, orden_compra, condicion_pago, "
        "vencimiento, fecha_factura, ingreso_id, item"
    )
    
    # Aplicar filtros de BD si existen
    if filtros_base:
        if filtros_base.get('proveedor'):
            query = query.ilike("proveedor_nombre", f"%{filtros_base['proveedor']}%")
        if filtros_base.get('proyecto'):
            try:
                query = query.eq("proyecto", int(filtros_base['proyecto']))
            except ValueError:
                pass
        if filtros_base.get('fecha_desde'):
            query = query.gte("fecha", filtros_base['fecha_desde'])
        if filtros_base.get('fecha_hasta'):
            query = query.lte("fecha", filtros_base['fecha_hasta'])
        if filtros_base.get('orden_numero'):
            try:
                query = query.eq("orden_numero", int(filtros_base['orden_numero']))
            except ValueError:
                pass
    
    query = query.order("orden_numero", desc=True)
    
    # Obtener todos los registros en lotes
    all_rows = []
    batch_size = 1000
    offset = 0
    
    while True:
        batch_result = query.range(offset, offset + batch_size - 1).execute()
        if not batch_result.data:
            break
        all_rows.extend(batch_result.data)
        if len(batch_result.data) < batch_size:
            break
        offset += batch_size
    
    if not all_rows:
        return []
    
    # Agrupar por orden_numero
    pagos_dict = {}
    orden_numeros = []
    
    for r in all_rows:
        num = r["orden_numero"]
        if num not in pagos_dict:
            orden_numeros.append(num)
            pagos_dict[num] = {
                "orden_numero": num,
                "fecha": r["fecha"],
                "proveedor": r.get("proveedor"),
                "proveedor_nombre": r["proveedor_nombre"],
                "detalle_compra": r["detalle_compra"],
                "factura": r["factura"],
                "total_pago": 0,
                "proyecto": r["proyecto"],
                "orden_compra": r["orden_compra"],
                "condicion_pago": r["condicion_pago"],
                "vencimiento": r["vencimiento"],
                "fecha_factura": r["fecha_factura"],
                "ingreso_id": r.get("ingreso_id"),
                "fecha_pago": None,
                "total_abonado": 0.0,
                "saldo_pendiente": 0.0,
                "item": r.get("item") or "-",  # ✅ FIX: Usar valor real de BD
                "rut_proveedor": "-",
                "fac_pago": ""
            }
        try:
            monto = int(round(float(r.get("costo_final_con_iva") or 0)))
        except Exception:
            monto = 0
        pagos_dict[num]["total_pago"] += monto
    
    pagos_list = list(pagos_dict.values())
    
    # Obtener datos relacionados EN LOTES
    if orden_numeros:
        batch_size = 100
        fecha_map = {}
        abonos_map = {}
        
        for i in range(0, len(orden_numeros), batch_size):
            batch = orden_numeros[i:i+batch_size]
            
            # Fechas de pago
            try:
                fechas_result = supabase.table("fechas_de_pagos_op").select(
                    "orden_numero, fecha_pago"
                ).in_("orden_numero", batch).execute()
                for r in (fechas_result.data or []):
                    fecha_map[r["orden_numero"]] = r["fecha_pago"]
            except Exception as e:
                logger.error(f"Error obteniendo fechas: {e}")
            
            # Abonos
            try:
                abonos_result = supabase.table("abonos_op").select(
                    "orden_numero, monto_abono"
                ).in_("orden_numero", batch).execute()
                for ab in (abonos_result.data or []):
                    num = ab["orden_numero"]
                    try:
                        monto_ab = int(round(float(ab.get("monto_abono") or 0)))
                    except Exception:
                        monto_ab = 0
                    abonos_map[num] = abonos_map.get(num, 0) + monto_ab
            except Exception as e:
                logger.error(f"Error obteniendo abonos: {e}")
        
        # Proveedores y proyectos (cache simple)
        proveedores = get_cached_proveedores()
        proyectos = get_cached_proyectos()
        rut_map = {p["id"]: p.get("rut", "-") for p in proveedores}
        proyecto_map = {p["id"]: p["proyecto"] for p in proyectos}
        
        # Aplicar datos a cada pago
        for pago in pagos_list:
            num = pago["orden_numero"]
            fecha_pago_bd = fecha_map.get(num)
            total_abonado = int(abonos_map.get(num, 0))
            total_pago = int(pago["total_pago"])
            saldo = max(0, total_pago - total_abonado)
            
            # Calcular estado
            estado = calcular_estado_pago(fecha_pago_bd, total_abonado, total_pago)
            
            # Determinar fecha a mostrar
            if fecha_pago_bd and total_abonado == 0:
                fecha_pago_mostrar = fecha_pago_bd
            elif total_abonado > 0 and saldo <= 0:
                fecha_pago_mostrar = fecha_pago_bd or date.today().isoformat()
            else:
                fecha_pago_mostrar = None
            
            pago["fecha_pago"] = fecha_pago_mostrar
            pago["total_abonado"] = total_abonado
            pago["saldo_pendiente"] = saldo
            pago["rut_proveedor"] = rut_map.get(pago.get("proveedor"), "-")
            pago["proyecto_nombre"] = proyecto_map.get(pago["proyecto"], f"Proyecto {pago['proyecto']}")
            pago["estado"] = estado
            
            # Formatear fechas a formato chileno usando función auxiliar
            pago["fecha"] = formatear_fecha_chilena(pago.get("fecha"))
            pago["vencimiento"] = formatear_fecha_chilena(pago.get("vencimiento"), "---")
    
    return pagos_list

# ================================================================
# ENDPOINT PRINCIPAL - LISTAR PAGOS CON PAGINACIÓN
# ================================================================

@bp.route("/list", methods=["GET"])
@token_required
def list_pagos(current_user):
    """
    Obtiene lista de pagos con paginación optimizada usando caché.
    """
    supabase = current_app.config['SUPABASE']
    
    try:
        # Parámetros de paginación
        page = max(1, request.args.get('page', 1, type=int))
        per_page = min(request.args.get('per_page', 50, type=int), 100)
        
        # Filtros
        proveedor = request.args.get('proveedor', '').strip()
        proyecto_id = request.args.get('proyecto', '').strip()
        estado_filtro = request.args.get('estado', '').strip()
        
        # Normalizar estado
        if estado_filtro:
            ef = estado_filtro.lower().strip()
            if ef in ("con abonos", "con_abonos", "con-abonos", "conabonos", "con abono", "con_abono"):
                estado_filtro = "abono"
            elif ef in ("con pagos", "pagadas", "pagado", "pagados", "pagado(s)"):
                estado_filtro = "pagado"
            elif ef in ("pendiente", "pendientes"):
                estado_filtro = "pendiente"
        
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        orden_numero = request.args.get('orden_numero', '').strip()
        
        # Calcular offset
        offset = (page - 1) * per_page
        
        # Crear key de caché basado en filtros de BD
        filtros_bd = {
            'proveedor': proveedor,
            'proyecto': proyecto_id,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'orden_numero': orden_numero
        }
        filtros_bd = {k: v for k, v in filtros_bd.items() if v}  # Solo los que tienen valor
        
        cache_key = str(sorted(filtros_bd.items()))
        current_time = time.time()
        
        # OPTIMIZACIÓN: Si hay filtro de estado, usar caché
        if estado_filtro:
            # Verificar si tenemos datos en caché y son recientes
            if (_pagos_cache["data"] is not None and 
                _pagos_cache.get("cache_key") == cache_key and
                current_time - _pagos_cache["timestamp"] < _pagos_cache["ttl"]):
                
                logger.info(f"✅ Usando caché de pagos (edad: {int(current_time - _pagos_cache['timestamp'])}s)")
                pagos_list = _pagos_cache["data"]
            else:
                # Obtener y procesar todos los pagos
                logger.info(f"🔄 Recargando caché de pagos...")
                start_time = time.time()
                pagos_list = obtener_todos_pagos_procesados(supabase, filtros_bd)
                elapsed = time.time() - start_time
                logger.info(f"✅ Caché recargado en {elapsed:.2f}s - {len(pagos_list)} órdenes procesadas")
                
                # Guardar en caché
                _pagos_cache["data"] = pagos_list
                _pagos_cache["cache_key"] = cache_key
                _pagos_cache["timestamp"] = current_time
            
            # Filtrar por estado
            pagos_filtrados = [p for p in pagos_list if p["estado"] == estado_filtro]
            total_filtrado = len(pagos_filtrados)
            
            # Aplicar paginación manual
            pagos_pagina = pagos_filtrados[offset:offset + per_page]
            total_pages = (total_filtrado + per_page - 1) // per_page
            
            return jsonify({
                "success": True,
                "data": {
                    "pagos": pagos_pagina,
                    "pagination": {
                        "page": page,
                        "per_page": per_page,
                        "total": total_filtrado,
                        "total_pages": total_pages
                    }
                }
            })
        
        # SIN filtro de estado: usar paginación normal de Supabase (más rápida)
        # Función auxiliar para reintentos
        def execute_with_retry(query_func, max_retries=3):
            for attempt in range(max_retries):
                try:
                    return query_func()
                except Exception as e:
                    if attempt == max_retries - 1:
                        raise
                    logger.warning(f"Intento {attempt + 1} falló: {e}. Reintentando...")
                    time.sleep(0.5 * (attempt + 1))
            return None
        
        # Calcular offset para paginación normal
        offset = (page - 1) * per_page
        end_range = offset + per_page - 1
        
        # Query base con paginación de Supabase
        def build_query():
            query = supabase.table("orden_de_pago").select(
                "orden_numero, fecha, proveedor, proveedor_nombre, detalle_compra, "
                "factura, costo_final_con_iva, proyecto, orden_compra, condicion_pago, "
                "vencimiento, fecha_factura, ingreso_id, item",  # ✅ FIX: Agregado item
                count="exact"
            )
            
            # Aplicar filtros de BD
            if proveedor:
                query = query.ilike("proveedor_nombre", f"%{proveedor}%")
            if proyecto_id:
                try:
                    query = query.eq("proyecto", int(proyecto_id))
                except ValueError:
                    pass
            if fecha_desde:
                query = query.gte("fecha", fecha_desde)
            if fecha_hasta:
                query = query.lte("fecha", fecha_hasta)
            if orden_numero:
                try:
                    query = query.eq("orden_numero", int(orden_numero))
                except ValueError:
                    pass
            
            # Ordenar y paginar
            query = query.order("orden_numero", desc=True).range(offset, end_range)
            return query.execute()
        
        result = execute_with_retry(build_query)
        
        if not result:
            return jsonify({
                "success": False,
                "message": "Error de conexión con la base de datos"
            }), 503
        
        # Procesar resultados (solo para queries sin filtro de estado)
        all_rows = result.data or []
        total_count = result.count or 0
        
        if not all_rows:
            return jsonify({
                "success": True,
                "data": {
                    "pagos": [],
                    "pagination": {
                        "page": page,
                        "per_page": per_page,
                        "total": 0,
                        "total_pages": 0
                    }
                }
            })
        
        # Agrupar por orden_numero
        pagos_dict = {}
        orden_numeros = []
        
        for r in all_rows:
            num = r["orden_numero"]
            if num not in pagos_dict:
                orden_numeros.append(num)
                pagos_dict[num] = {
                    "orden_numero": num,
                    "fecha": r["fecha"],
                    "proveedor": r.get("proveedor"),
                    "proveedor_nombre": r["proveedor_nombre"],
                    "detalle_compra": r["detalle_compra"],
                    "factura": r["factura"],
                    "total_pago": 0,
                    "proyecto": r["proyecto"],
                    "orden_compra": r["orden_compra"],
                    "condicion_pago": r["condicion_pago"],
                    "vencimiento": r["vencimiento"],
                    "fecha_factura": r["fecha_factura"],
                    "ingreso_id": r.get("ingreso_id"),
                    "fecha_pago": None,
                    "total_abonado": 0.0,
                    "saldo_pendiente": 0.0,
                    "item": r.get("item") or "-",  # ✅ FIX: Usar valor real de BD
                    "rut_proveedor": "-",
                    "fac_pago": ""
                }
            try:
                monto = int(round(float(r.get("costo_final_con_iva") or 0)))
            except Exception:
                monto = 0
            pagos_dict[num]["total_pago"] += monto
        
        pagos_list = list(pagos_dict.values())
        
        # Obtener datos relacionados en lotes
        if orden_numeros:
            batch_size = 50
            fecha_map = {}
            abonos_map = {}
            
            for i in range(0, len(orden_numeros), batch_size):
                batch = orden_numeros[i:i+batch_size]
                
                # Fechas de pago
                try:
                    fechas_result = execute_with_retry(
                        lambda: supabase.table("fechas_de_pagos_op").select(
                            "orden_numero, fecha_pago"
                        ).in_("orden_numero", batch).execute()
                    )
                    for r in (fechas_result.data or []):
                        fecha_map[r["orden_numero"]] = r["fecha_pago"]
                except Exception as e:
                    logger.error(f"Error obteniendo fechas: {e}")
                
                # Abonos
                try:
                    abonos_result = execute_with_retry(
                        lambda: supabase.table("abonos_op").select(
                            "orden_numero, monto_abono"
                        ).in_("orden_numero", batch).execute()
                    )
                    for ab in (abonos_result.data or []):
                        num = ab["orden_numero"]
                        try:
                            monto_ab = int(round(float(ab.get("monto_abono") or 0)))
                        except Exception:
                            monto_ab = 0
                        abonos_map[num] = abonos_map.get(num, 0) + monto_ab
                except Exception as e:
                    logger.error(f"Error obteniendo abonos: {e}")
            
            # Proveedores y proyectos
            proveedores = get_cached_proveedores()
            proyectos = get_cached_proyectos()
            rut_map = {p["id"]: p.get("rut", "-") for p in proveedores}
            proyecto_map = {p["id"]: p["proyecto"] for p in proyectos}
            
            # Aplicar datos a cada pago
            for pago in pagos_list:
                num = pago["orden_numero"]
                fecha_pago_bd = fecha_map.get(num)
                total_abonado = int(abonos_map.get(num, 0))
                total_pago = int(pago["total_pago"])
                saldo = max(0, total_pago - total_abonado)
                
                estado = calcular_estado_pago(fecha_pago_bd, total_abonado, total_pago)
                
                # Fecha a mostrar
                if fecha_pago_bd and total_abonado == 0:
                    fecha_pago_mostrar = fecha_pago_bd
                elif total_abonado > 0 and saldo <= 0:
                    fecha_pago_mostrar = fecha_pago_bd or date.today().isoformat()
                else:
                    fecha_pago_mostrar = None
                
                pago["fecha_pago"] = fecha_pago_mostrar
                pago["total_abonado"] = total_abonado
                pago["saldo_pendiente"] = saldo
                pago["rut_proveedor"] = rut_map.get(pago.get("proveedor"), "-")
                pago["proyecto_nombre"] = proyecto_map.get(pago["proyecto"], f"Proyecto {pago['proyecto']}")
                pago["estado"] = estado
                
                # Formatear fechas a formato chileno usando función auxiliar
                pago["fecha"] = formatear_fecha_chilena(pago.get("fecha"))
                pago["vencimiento"] = formatear_fecha_chilena(pago.get("vencimiento"), "---")
        
        # Paginación
        total_pages = (total_count + per_page - 1) // per_page
        
        return jsonify({
            "success": True,
            "data": {
                "pagos": pagos_list,
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total": total_count,
                    "total_pages": total_pages
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Error en list_pagos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False, 
            "message": "Error interno del servidor. Intente nuevamente."
        }), 500

# ================================================================
# ENDPOINT - ESTADÍSTICAS
# ================================================================

@bp.route("/stats", methods=["GET"])
@token_required
def get_stats(current_user):
    """
    Obtiene estadísticas generales de pagos con retry logic.
    Aplica los mismos filtros que list_pagos.
    """
    supabase = current_app.config['SUPABASE']
    
    MAX_RETRIES = 3
    RETRY_DELAY = 0.5
    
    for intento in range(MAX_RETRIES):
        try:
            # Filtros (mismos que list_pagos)
            proveedor = request.args.get('proveedor', '').strip()
            proyecto_id = request.args.get('proyecto', '').strip()
            fecha_desde = request.args.get('fecha_desde', '').strip()
            fecha_hasta = request.args.get('fecha_hasta', '').strip()
            
            # Obtener todas las filas de `orden_de_pago` aplicando los mismos filtros
            # que usa la vista detallada. Para evitar límites de Supabase, hacemos
            # paginado por lotes (batching) igual que el sistema antiguo.
            page_size = 1000
            offset = 0
            all_rows = []

            while True:
                q = supabase.table("orden_de_pago").select(
                    "orden_numero, costo_final_con_iva, proyecto"
                ).order("orden_numero", desc=True).range(offset, offset + page_size - 1)

                if proveedor:
                    q = q.ilike("proveedor_nombre", f"%{proveedor}%")
                if proyecto_id:
                    try:
                        q = q.eq("proyecto", int(proyecto_id))
                    except ValueError:
                        pass
                if fecha_desde:
                    q = q.gte("fecha", fecha_desde)
                if fecha_hasta:
                    q = q.lte("fecha", fecha_hasta)

                res = q.execute()
                batch = res.data or []
                all_rows.extend(batch)
                if len(batch) < page_size:
                    break
                offset += page_size
            
            # Agrupar por orden_numero
            pagos_dict = {}
            for r in all_rows:
                num = r["orden_numero"]
                if num not in pagos_dict:
                    pagos_dict[num] = {
                        "orden_numero": num,
                        "total_pago": 0,
                        "proyecto": r["proyecto"]
                    }
                try:
                    monto = int(round(float(r.get("costo_final_con_iva") or 0)))
                except Exception:
                    monto = 0
                pagos_dict[num]["total_pago"] += monto
            
            orden_numeros = list(pagos_dict.keys())
            
            # Obtener fechas de pago y abonos
            fecha_map = {}
            abonos_map = {}

            # Para evitar problemas de tipos y asegurar coincidencia con la vista
            # antigua, traemos las filas de `fechas_de_pagos_op` por lotes y construimos
            # un mapa con claves como strings y como ints.
            try:
                page_size = 1000
                off = 0
                fechas_data = []
                while True:
                    batch = supabase.table("fechas_de_pagos_op").select("orden_numero, fecha_pago").range(off, off + page_size - 1).execute().data or []
                    if not batch:
                        break
                    fechas_data.extend(batch)
                    off += page_size
                # Construir mapa con claves string e int
                for row in fechas_data:
                    k = row.get("orden_numero")
                    v = row.get("fecha_pago")
                    if k is None:
                        continue
                    fecha_map[str(k)] = v
                    try:
                        fecha_map[int(k)] = v
                    except Exception:
                        pass
            except Exception as e:
                logger.error(f"Error obteniendo fechas en batches: {e}")

            # Abonos: traer todos los abonos y acumular por orden
            try:
                abonos_data = supabase.table("abonos_op").select("orden_numero, monto_abono").execute().data or []
                for ab in abonos_data:
                    num = ab.get("orden_numero")
                    try:
                        monto = int(round(float(ab.get("monto_abono") or 0)))
                    except Exception:
                        monto = 0
                    if num is None:
                        continue
                    # Asegurar que usamos int como clave única
                    try:
                        num_int = int(num)
                        abonos_map[num_int] = abonos_map.get(num_int, 0) + monto
                    except Exception:
                        # Si no se puede convertir a int, usar el valor original
                        abonos_map[num] = abonos_map.get(num, 0) + monto
            except Exception as e:
                logger.error(f"Error obteniendo abonos: {e}")
            
            # Calcular métricas siguiendo la nueva regla de negocio:
            # - Se parte de los valores únicos de `orden_numero` en `orden_de_pago` (ya está en pagos_dict)
            # - Si una orden aparece en `fechas_de_pagos_op` la consideramos pagada y NO la includimos
            #   en el cálculo de "total_pendiente".
            total_ordenes = len(pagos_dict)

            # Conjunto de órdenes con fecha de pago (pagadas)
            ordenes_con_fecha = set(fecha_map.keys())

            # Contadores
            pagadas = 0
            pendientes = 0
            con_abonos = 0

            total_pendiente = 0.0
            total_saldo_abonos = 0.0
            saldo_por_proyecto = {}
            monto_total_por_proyecto = {}  # Nuevo: suma de todas las OP por proyecto

            # Proyectos
            proyectos = get_cached_proyectos()
            proyecto_map = {p["id"]: p["proyecto"] for p in proyectos}

            # Recorremos las órdenes únicas y separamos los montos:
            # - "Pendientes" = total_pago de órdenes SIN abonos ni fecha_pago
            # - "Saldo Abonos" = SALDO (total - abonos) de órdenes con abonos SIN fecha_pago
            
            # Para debug: registrar todas las órdenes con abonos
            ordenes_con_abonos_detalle = []
            
            for num, pago in pagos_dict.items():
                total_abonado = abonos_map.get(num, 0)
                total_pago = pago["total_pago"]
                fecha_pago = fecha_map.get(num)
                proyecto_nombre = proyecto_map.get(pago.get("proyecto"), str(pago.get("proyecto")))

                # Sumar SIEMPRE al monto total del proyecto (todas las OP)
                monto_total_por_proyecto[proyecto_nombre] = monto_total_por_proyecto.get(proyecto_nombre, 0) + total_pago

                # PRIMERO: Si tiene abonos Y NO tiene fecha_pago, calcular saldo y sumar
                if total_abonado > 0 and not fecha_pago:
                    con_abonos += 1
                    saldo = max(0, total_pago - total_abonado)
                    total_saldo_abonos += saldo
                    saldo_por_proyecto[proyecto_nombre] = saldo_por_proyecto.get(proyecto_nombre, 0) + saldo
                    
                    # Debug: guardar detalle
                    ordenes_con_abonos_detalle.append({
                        "orden": num,
                        "total": total_pago,
                        "abonado": total_abonado,
                        "saldo": saldo,
                        "fecha_pago": fecha_pago,
                        "proyecto": proyecto_nombre
                    })

                # SEGUNDO: Calcular estado para contadores (pagado vs pendiente)
                estado = calcular_estado_pago(fecha_pago, total_abonado, total_pago)

                if estado == "pagado":
                    pagadas += 1
                elif estado == "pendiente":
                    # Pendiente sin abonos - sumar el total completo a "Pendientes"
                    pendientes += 1
                    total_pendiente += total_pago
                    saldo_por_proyecto[proyecto_nombre] = saldo_por_proyecto.get(proyecto_nombre, 0) + total_pago

            # Total general
            total_general_pendiente = total_pendiente + total_saldo_abonos

            # Log de depuración para diagnosticar valores inesperados
            logger.debug(f"get_stats: filas_consultadas={len(all_rows)}, ordenes_unicas={len(pagos_dict)}, ordenes_con_fecha={len(ordenes_con_fecha)}")

            # Filtrar proyectos con saldo significativo (> 1 para evitar ruidos)
            saldo_por_proyecto = {k: v for k, v in saldo_por_proyecto.items() if v > 1}
            
            # Filtrar monto_total solo para proyectos que tienen saldo pendiente
            monto_total_por_proyecto_filtrado = {
                k: monto_total_por_proyecto.get(k, 0) 
                for k in saldo_por_proyecto.keys()
            }

            logger.info(f"📊 Stats: pendientes={pendientes}, con_abonos={con_abonos}, monto_pendiente={total_pendiente}, saldo_abonos={total_saldo_abonos}")

            # Si se solicita debug, devolver datos adicionales para diagnóstico
            if request.args.get('debug') == '1':
                muestra_ordenes = list(pagos_dict.keys())[:20]
                muestra_fechas = {k: fecha_map.get(k) for k in muestra_ordenes}
                muestra_abonos = {k: abonos_map.get(k, 0) for k in muestra_ordenes}
                return jsonify({
                    "success": True,
                    "data": {
                        "total_ordenes": total_ordenes,
                        "pagadas": pagadas,
                        "pendientes": pendientes,
                        "con_abonos": con_abonos,
                        "monto_pendiente": total_pendiente,
                        "saldo_abonos": total_saldo_abonos,
                        "total_general": total_general_pendiente,
                        "saldo_por_proyecto": saldo_por_proyecto,
                        "monto_total_por_proyecto": monto_total_por_proyecto_filtrado,
                        "debug": {
                            "filas_consultadas": len(all_rows),
                            "ordenes_unicas_consideradas": len(pagos_dict),
                            "ordenes_con_fecha_count": len(fecha_map),
                            "ordenes_con_abonos_count": len(abonos_map),
                            "muestra_ordenes": muestra_ordenes,
                            "muestra_fechas": muestra_fechas,
                            "muestra_abonos": muestra_abonos,
                            "ordenes_con_abonos_detalle": ordenes_con_abonos_detalle
                        }
                    }
                })

            return jsonify({
                "success": True,
                "data": {
                    "total_ordenes": total_ordenes,
                    "pagadas": pagadas,
                    "pendientes": pendientes,
                    "con_abonos": con_abonos,
                    "monto_pendiente": total_pendiente,
                    "saldo_abonos": total_saldo_abonos,
                    "total_general": total_general_pendiente,
                    "saldo_por_proyecto": saldo_por_proyecto,
                    "monto_total_por_proyecto": monto_total_por_proyecto_filtrado
                }
            })
            
        except Exception as e:
            if intento < MAX_RETRIES - 1:
                logger.warning(f"Intento {intento + 1} falló en get_stats: {e}. Reintentando...")
                import time
                time.sleep(RETRY_DELAY)
                continue
            else:
                logger.error(f"Error en get_stats después de {MAX_RETRIES} intentos: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({"success": False, "message": "Error interno del servidor"}), 500
    
    # Si llegamos aquí es porque todos los reintentos fallaron
    return jsonify({"success": False, "message": "Error interno del servidor"}), 500

# ================================================================
# ENDPOINT - ACTUALIZAR FECHA DE PAGO
# ================================================================

@bp.route("/fecha", methods=["PUT"])
@token_required
def update_fecha_pago(current_user):
    """
    Actualiza la fecha de pago de una orden.
    Body: { "orden_numero": 123, "fecha_pago": "2025-10-24" }
    """
    supabase = current_app.config['SUPABASE']
    
    try:
        data = request.get_json()
        orden_numero = data.get("orden_numero")
        fecha_pago = data.get("fecha_pago", "").strip()
        
        if not orden_numero:
            return jsonify({
                "success": False, 
                "message": "orden_numero requerido"
            }), 400
        
        # Validar fecha
        if fecha_pago:
            try:
                fecha_obj = datetime.strptime(fecha_pago, "%Y-%m-%d").date()
                if fecha_obj > date.today() + timedelta(days=7):
                    return jsonify({
                        "success": False,
                        "message": "Fecha de pago muy futura"
                    }), 400
            except ValueError:
                return jsonify({
                    "success": False,
                    "message": "Formato de fecha inválido"
                }), 400
            
            # Upsert fecha
            supabase.table("fechas_de_pagos_op").upsert({
                "orden_numero": orden_numero,
                "fecha_pago": fecha_pago
            }, on_conflict=["orden_numero"]).execute()
            
            # Invalidar caché
            invalidar_cache_pagos()
            
            return jsonify({
                "success": True,
                "message": "Fecha de pago actualizada"
            })
        else:
            # Eliminar fecha si está vacía (verificar abonos primero)
            abonos = supabase.table("abonos_op").select("id").eq(
                "orden_numero", orden_numero
            ).execute()
            
            if abonos.data:
                return jsonify({
                    "success": False,
                    "message": "No puedes borrar la fecha porque tiene abonos registrados"
                }), 400
            
            # Eliminar fecha
            supabase.table("fechas_de_pagos_op").delete().eq(
                "orden_numero", orden_numero
            ).execute()
            
            # Invalidar caché
            invalidar_cache_pagos()
            
            return jsonify({
                "success": True,
                "message": "Fecha de pago eliminada"
            })
        
    except Exception as e:
        logger.error(f"Error en update_fecha_pago: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ================================================================
# ENDPOINT - OBTENER ABONOS DE UNA ORDEN
# ================================================================

@bp.route("/abonos/<int:orden_numero>", methods=["GET"])
@token_required
def get_abonos(current_user, orden_numero):
    """Obtiene todos los abonos de una orden"""
    supabase = current_app.config['SUPABASE']
    
    try:
        result = supabase.table("abonos_op").select(
            "id, orden_numero, monto_abono, fecha_abono, observacion"
        ).eq("orden_numero", orden_numero).order("fecha_abono", desc=True).execute()
        
        return jsonify({
            "success": True,
            "data": result.data or []
        })
        
    except Exception as e:
        logger.error(f"Error en get_abonos: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ================================================================
# ENDPOINT - CREAR ABONO
# ================================================================

@bp.route("/abono", methods=["POST"])
@token_required
def create_abono(current_user):
    """
    Crea un nuevo abono.
    Body: {
        "orden_numero": 123,
        "monto_abono": 50000,
        "fecha_abono": "2025-10-24",
        "observacion": "Abono parcial"
    }
    """
    supabase = current_app.config['SUPABASE']
    
    try:
        data = request.get_json()
        orden_numero = data.get("orden_numero")
        monto_abono = data.get("monto_abono")
        fecha_abono = data.get("fecha_abono")
        observacion = data.get("observacion", "")
        
        if not all([orden_numero, monto_abono, fecha_abono]):
            return jsonify({
                "success": False,
                "message": "Faltan datos requeridos"
            }), 400
        
        # Validar monto
        monto_abono = int(round(float(monto_abono)))
        if monto_abono <= 0:
            return jsonify({
                "success": False,
                "message": "El monto debe ser mayor a cero"
            }), 400
        
        # Obtener total de la orden (suma de todas las líneas)
        orden = supabase.table("orden_de_pago").select(
            "costo_final_con_iva"
        ).eq("orden_numero", orden_numero).execute()  # ✅ FIX: Quitar limit(1) para obtener todas las líneas
        
        if not orden.data:
            return jsonify({
                "success": False,
                "message": "Orden no encontrada"
            }), 404
        
        # ✅ FIX: Sumar todas las líneas de la orden
        total_pago = sum(
            int(round(float(linea.get("costo_final_con_iva") or 0)))
            for linea in orden.data
        )
        
        # Obtener suma de abonos existentes
        abonos = supabase.table("abonos_op").select("monto_abono").eq(
            "orden_numero", orden_numero
        ).execute()
        
        suma_abonos = sum(
            int(round(float(a.get("monto_abono") or 0))) 
            for a in (abonos.data or [])
        )
        
        # Validar que no exceda el total
        nueva_suma = suma_abonos + monto_abono
        if nueva_suma > total_pago:
            return jsonify({
                "success": False,
                "message": f"La suma de abonos ({nueva_suma}) supera el total ({total_pago})"
            }), 400
        
        # Insertar abono
        result = supabase.table("abonos_op").insert({
            "orden_numero": orden_numero,
            "monto_abono": monto_abono,
            "fecha_abono": fecha_abono,
            "observacion": observacion
        }).execute()
        
        # Si completa el total, asignar fecha de pago
        if nueva_suma == total_pago:
            supabase.table("fechas_de_pagos_op").upsert({
                "orden_numero": orden_numero,
                "fecha_pago": fecha_abono
            }, on_conflict=["orden_numero"]).execute()
        
        # Invalidar caché
        invalidar_cache_pagos()
        
        return jsonify({
            "success": True,
            "message": "Abono registrado exitosamente",
            "data": result.data[0] if result.data else None
        })
        
    except Exception as e:
        logger.error(f"Error en create_abono: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ================================================================
# ENDPOINT - EDITAR ABONO
# ================================================================

@bp.route("/abono/<int:abono_id>", methods=["PUT"])
@token_required
def update_abono(current_user, abono_id):
    """Edita un abono existente"""
    supabase = current_app.config['SUPABASE']
    
    try:
        data = request.get_json()
        monto_abono = int(round(float(data.get("monto_abono") or 0)))
        fecha_abono = data.get("fecha_abono")
        observacion = data.get("observacion")
        
        if monto_abono <= 0:
            return jsonify({
                "success": False,
                "message": "El monto debe ser mayor a cero"
            }), 400
        
        # Obtener orden_numero del abono
        abono = supabase.table("abonos_op").select("orden_numero").eq(
            "id", abono_id
        ).limit(1).execute()
        
        if not abono.data:
            return jsonify({
                "success": False,
                "message": "Abono no encontrado"
            }), 404
        
        orden_numero = abono.data[0]["orden_numero"]
        
        # Obtener total de la orden
        orden = supabase.table("orden_de_pago").select(
            "costo_final_con_iva"
        ).eq("orden_numero", orden_numero).limit(1).execute()
        
        total_pago = int(round(float(orden.data[0].get("costo_final_con_iva") or 0)))
        
        # Suma de otros abonos (excluyendo el actual)
        otros_abonos = supabase.table("abonos_op").select("monto_abono").eq(
            "orden_numero", orden_numero
        ).neq("id", abono_id).execute()
        
        suma_otros = sum(
            int(round(float(a.get("monto_abono") or 0))) 
            for a in (otros_abonos.data or [])
        )
        
        # Validar nueva suma
        nueva_suma = suma_otros + monto_abono
        if nueva_suma > total_pago:
            return jsonify({
                "success": False,
                "message": f"La suma de abonos ({nueva_suma}) supera el total ({total_pago})"
            }), 400
        
        # Actualizar abono
        update_data = {"monto_abono": monto_abono}
        if fecha_abono:
            update_data["fecha_abono"] = fecha_abono
        if observacion is not None:
            update_data["observacion"] = observacion
        
        supabase.table("abonos_op").update(update_data).eq("id", abono_id).execute()
        
        # Actualizar fecha de pago si completa el total
        if nueva_suma == total_pago:
            supabase.table("fechas_de_pagos_op").upsert({
                "orden_numero": orden_numero,
                "fecha_pago": fecha_abono or date.today().isoformat()
            }, on_conflict=["orden_numero"]).execute()
        elif nueva_suma < total_pago:
            # Si no completa, eliminar fecha automática (solo si fue puesta por abonos)
            pass
        
        # Invalidar caché
        invalidar_cache_pagos()
        
        return jsonify({
            "success": True,
            "message": "Abono actualizado exitosamente"
        })
        
    except Exception as e:
        logger.error(f"Error en update_abono: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ================================================================
# ENDPOINT - ELIMINAR ABONO
# ================================================================

@bp.route("/abono/<int:abono_id>", methods=["DELETE"])
@token_required
def delete_abono(current_user, abono_id):
    """Elimina un abono"""
    supabase = current_app.config['SUPABASE']
    
    try:
        # Obtener orden_numero del abono
        abono = supabase.table("abonos_op").select("orden_numero").eq(
            "id", abono_id
        ).limit(1).execute()
        
        if not abono.data:
            return jsonify({
                "success": False,
                "message": "Abono no encontrado"
            }), 404
        
        orden_numero = abono.data[0]["orden_numero"]
        
        # Eliminar abono
        supabase.table("abonos_op").delete().eq("id", abono_id).execute()
        
        # Verificar abonos restantes
        abonos_restantes = supabase.table("abonos_op").select("monto_abono").eq(
            "orden_numero", orden_numero
        ).execute()
        
        suma_restante = sum(
            int(round(float(a.get("monto_abono") or 0))) 
            for a in (abonos_restantes.data or [])
        )
        
        # Si no quedan abonos o no completan el total, eliminar fecha automática
        if suma_restante == 0:
            supabase.table("fechas_de_pagos_op").delete().eq(
                "orden_numero", orden_numero
            ).execute()
        else:
            # Verificar si completa el total
            orden = supabase.table("orden_de_pago").select(
                "costo_final_con_iva"
            ).eq("orden_numero", orden_numero).limit(1).execute()
            
            total_pago = int(round(float(orden.data[0].get("costo_final_con_iva") or 0)))
            
            if suma_restante < total_pago:
                supabase.table("fechas_de_pagos_op").delete().eq(
                    "orden_numero", orden_numero
                ).execute()
        
        # Invalidar caché
        invalidar_cache_pagos()
        
        return jsonify({
            "success": True,
            "message": "Abono eliminado exitosamente"
        })
        
    except Exception as e:
        logger.error(f"Error en delete_abono: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ================================================================
# ENDPOINT - LISTAS PARA FILTROS
# ================================================================

@bp.route("/filters", methods=["GET"])
@token_required
def get_filters(current_user):
    """Obtiene listas para los filtros (proveedores, proyectos)"""
    try:
        proveedores = get_cached_proveedores()
        proyectos = get_cached_proyectos()
        
        return jsonify({
            "success": True,
            "data": {
                "proveedores": [
                    {"id": p["id"], "nombre": p["nombre"]} 
                    for p in proveedores
                ],
                "proyectos": [
                    {"id": p["id"], "nombre": p["proyecto"]} 
                    for p in proyectos
                ]
            }
        })
        
    except Exception as e:
        logger.error(f"Error en filters: {e}")
        return jsonify({"success": False, "message": "Error al cargar filtros"}), 500


# ================================================================
# EXPORTAR A EXCEL
# ================================================================

@bp.route("/exportar-excel", methods=["GET"])
@token_required
def exportar_pagos_excel(current_user):
    """
    Exporta las órdenes de pago a Excel aplicando los mismos filtros que la vista
    """
    try:
        supabase = current_app.config['SUPABASE']
        
        # Obtener filtros de los query params
        proveedor = request.args.get('proveedor', '').strip()
        proyecto_id = request.args.get('proyecto', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        estado = request.args.get('estado', '').strip()
        
        logger.info(f"Exportando Excel con filtros: proveedor={proveedor}, proyecto={proyecto_id}, estado={estado}")
        
        # Construir query con filtros
        query = supabase.table("orden_de_pago").select(
            "orden_numero, fecha, proveedor, proveedor_nombre, detalle_compra, "
            "factura, costo_final_con_iva, proyecto, item, orden_compra, vencimiento"
        ).order("orden_numero", desc=True)
        
        if proveedor:
            query = query.ilike("proveedor_nombre", f"%{proveedor}%")
        if proyecto_id:
            try:
                query = query.eq("proyecto", int(proyecto_id))
            except ValueError:
                pass
        if fecha_desde:
            query = query.gte("fecha", fecha_desde)
        if fecha_hasta:
            query = query.lte("fecha", fecha_hasta)
        
        # Ejecutar query en lotes para obtener TODOS los registros (sin límite de paginación)
        all_rows = []
        batch_size = 1000
        offset = 0
        
        while True:
            # Supabase requiere re-construir el query con .range() en cada iteración
            batch_result = supabase.table("orden_de_pago").select(
                "orden_numero, fecha, proveedor, proveedor_nombre, detalle_compra, "
                "factura, costo_final_con_iva, proyecto, item, orden_compra, vencimiento"
            ).order("orden_numero", desc=True)
            
            # Reaplicar filtros
            if proveedor:
                batch_result = batch_result.ilike("proveedor_nombre", f"%{proveedor}%")
            if proyecto_id:
                try:
                    batch_result = batch_result.eq("proyecto", int(proyecto_id))
                except ValueError:
                    pass
            if fecha_desde:
                batch_result = batch_result.gte("fecha", fecha_desde)
            if fecha_hasta:
                batch_result = batch_result.lte("fecha", fecha_hasta)
            
            # Aplicar paginación
            batch_result = batch_result.range(offset, offset + batch_size - 1)
            batch_data = batch_result.execute().data or []
            
            if not batch_data:
                break
                
            all_rows.extend(batch_data)
            
            # Si trajo menos registros que el batch_size, ya no hay más
            if len(batch_data) < batch_size:
                break
                
            offset += batch_size
        
        if not all_rows:
            return jsonify({"success": False, "message": "No hay datos para exportar"}), 404
        
        # Agrupar por orden_numero y calcular totales
        pagos_dict = {}
        for r in all_rows:
            num = r["orden_numero"]
            if num not in pagos_dict:
                pagos_dict[num] = {
                    "orden_numero": num,
                    "fecha": r["fecha"],
                    "proveedor": r.get("proveedor"),  # ID del proveedor
                    "proveedor_nombre": r["proveedor_nombre"],
                    "detalle_compra": r["detalle_compra"],
                    "factura": r["factura"],
                    "total_pago": 0,
                    "proyecto": r["proyecto"],
                    "item": r["item"],
                    "orden_compra": r["orden_compra"],
                    "vencimiento": r.get("vencimiento")
                }
            try:
                monto = int(round(float(r.get("costo_final_con_iva") or 0)))
            except Exception:
                monto = 0
            pagos_dict[num]["total_pago"] += monto
        
        # Obtener lista de orden_numeros únicos para consultas relacionadas
        orden_numeros = list(pagos_dict.keys())
        
        # Obtener fechas de pago en lotes (igual que en el listado)
        fecha_map = {}
        if orden_numeros:
            batch_size = 100
            for i in range(0, len(orden_numeros), batch_size):
                batch = orden_numeros[i:i+batch_size]
                try:
                    fechas_result = supabase.table("fechas_de_pagos_op").select(
                        "orden_numero, fecha_pago"
                    ).in_("orden_numero", batch).execute()
                    for r in (fechas_result.data or []):
                        fecha_map[r["orden_numero"]] = r["fecha_pago"]
                except Exception as e:
                    logger.error(f"Error obteniendo fechas: {e}")
        
        # Obtener abonos en lotes (igual que en el listado)
        abonos_map = {}
        if orden_numeros:
            batch_size = 100
            for i in range(0, len(orden_numeros), batch_size):
                batch = orden_numeros[i:i+batch_size]
                try:
                    abonos_result = supabase.table("abonos_op").select(
                        "orden_numero, monto_abono"
                    ).in_("orden_numero", batch).execute()
                    for ab in (abonos_result.data or []):
                        num = ab["orden_numero"]
                        try:
                            monto = int(round(float(ab.get("monto_abono") or 0)))
                        except Exception:
                            monto = 0
                        abonos_map[num] = abonos_map.get(num, 0) + monto
                except Exception as e:
                    logger.error(f"Error obteniendo abonos: {e}")
        
        # Obtener proyectos
        proyectos = get_cached_proyectos()
        proyecto_map = {p["id"]: p["proyecto"] for p in proyectos}
        
        # Obtener proveedores para obtener RUTs
        proveedores = get_cached_proveedores()
        rut_map = {p["id"]: p.get("rut", "-") for p in proveedores}
        
        logger.info(f"Procesando {len(pagos_dict)} órdenes únicas")
        logger.info(f"Fechas de pago encontradas: {len(fecha_map)}")
        logger.info(f"Órdenes con abonos: {len(abonos_map)}")
        
        # Calcular estados y saldos
        pagos_list = []
        estados_count = {"pagado": 0, "pendiente": 0, "abono": 0}
        
        for num, pago in pagos_dict.items():
            total_abonado = abonos_map.get(num, 0)
            total_pago = pago["total_pago"]
            fecha_pago = fecha_map.get(num)
            
            # Calcular estado usando la misma función
            estado_calculado = calcular_estado_pago(fecha_pago, total_abonado, total_pago)
            estados_count[estado_calculado] = estados_count.get(estado_calculado, 0) + 1
            
            # Filtrar por estado si se especificó
            if estado and estado_calculado != estado:
                continue
            
            saldo = 0 if fecha_pago else max(0, total_pago - total_abonado)
            proyecto_nombre = proyecto_map.get(pago.get("proyecto"), str(pago.get("proyecto")))
            
            # Obtener RUT desde el mapa de proveedores
            rut_proveedor = rut_map.get(pago.get("proveedor"), "-")
            
            # Formatear fechas usando función auxiliar
            fecha_formateada = formatear_fecha_chilena(pago["fecha"])
            vencimiento_formateado = formatear_fecha_chilena(pago.get("vencimiento"), "---")
            
            pagos_list.append({
                "orden_numero": num,
                "fecha": fecha_formateada,
                "vencimiento": vencimiento_formateado,
                "proveedor_nombre": pago["proveedor_nombre"],
                "rut_proveedor": rut_proveedor,
                "detalle_compra": pago["detalle_compra"],
                "factura": pago["factura"],
                "total_pago": total_pago,
                "proyecto_nombre": proyecto_nombre,
                "item": pago["item"],
                "orden_compra": pago["orden_compra"],
                "fecha_pago": fecha_pago or "",
                "total_abonado": total_abonado,
                "saldo_pendiente": saldo,
                "estado": estado_calculado
            })
        
        logger.info(f"Estados calculados - Pagado: {estados_count.get('pagado', 0)}, Pendiente: {estados_count.get('pendiente', 0)}, Abono: {estados_count.get('abono', 0)}")
        logger.info(f"Total de órdenes a exportar después de filtros: {len(pagos_list)}")
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Órdenes de Pago"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"INFORME DE ÓRDENES DE PAGO - {datetime.now().strftime('%d/%m/%Y')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Headers
        headers = ["OP", "Fecha", "Vencimiento", "Proveedor", "RUT", "Detalle", "Factura", "Total", 
                   "Proyecto", "Item", "OC", "Fecha Pago", "Abonos", "Saldo", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for idx, pago in enumerate(pagos_list, 4):
            ws.cell(row=idx, column=1, value=pago["orden_numero"]).border = border
            ws.cell(row=idx, column=2, value=pago["fecha"]).border = border
            ws.cell(row=idx, column=3, value=pago.get("vencimiento") or "---").border = border
            ws.cell(row=idx, column=4, value=pago["proveedor_nombre"]).border = border
            ws.cell(row=idx, column=5, value=pago["rut_proveedor"]).border = border
            ws.cell(row=idx, column=6, value=pago["detalle_compra"]).border = border
            ws.cell(row=idx, column=7, value=pago["factura"]).border = border
            
            # Total (formato moneda)
            cell_total = ws.cell(row=idx, column=8, value=pago["total_pago"])
            cell_total.number_format = '#,##0'
            cell_total.border = border
            
            ws.cell(row=idx, column=9, value=pago["proyecto_nombre"]).border = border
            ws.cell(row=idx, column=10, value=pago["item"]).border = border
            ws.cell(row=idx, column=11, value=pago["orden_compra"]).border = border
            ws.cell(row=idx, column=12, value=pago["fecha_pago"]).border = border
            
            # Abonos (formato moneda)
            cell_abonos = ws.cell(row=idx, column=13, value=pago["total_abonado"])
            cell_abonos.number_format = '#,##0'
            cell_abonos.border = border
            
            # Saldo (formato moneda)
            cell_saldo = ws.cell(row=idx, column=14, value=pago["saldo_pendiente"])
            cell_saldo.number_format = '#,##0'
            cell_saldo.border = border
            
            # Estado con color
            cell_estado = ws.cell(row=idx, column=15, value=pago["estado"].upper())
            cell_estado.border = border
            if pago["estado"] == "pagado":
                cell_estado.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                cell_estado.font = Font(color="065F46", bold=True)
            elif pago["estado"] == "pendiente":
                cell_estado.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                cell_estado.font = Font(color="991B1B", bold=True)
            else:  # abono
                cell_estado.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                cell_estado.font = Font(color="92400E", bold=True)
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 8   # OP
        ws.column_dimensions['B'].width = 12  # Fecha
        ws.column_dimensions['C'].width = 12  # Vencimiento
        ws.column_dimensions['D'].width = 30  # Proveedor
        ws.column_dimensions['E'].width = 15  # RUT
        ws.column_dimensions['F'].width = 40  # Detalle
        ws.column_dimensions['G'].width = 15  # Factura
        ws.column_dimensions['H'].width = 15  # Total
        ws.column_dimensions['I'].width = 25  # Proyecto
        ws.column_dimensions['J'].width = 10  # Item
        ws.column_dimensions['K'].width = 10  # OC
        ws.column_dimensions['L'].width = 12  # Fecha Pago
        ws.column_dimensions['M'].width = 15  # Abonos
        ws.column_dimensions['N'].width = 15  # Saldo
        ws.column_dimensions['O'].width = 12  # Estado
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"ordenes_pago_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error al exportar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error al generar Excel: {str(e)}"
        }), 500
